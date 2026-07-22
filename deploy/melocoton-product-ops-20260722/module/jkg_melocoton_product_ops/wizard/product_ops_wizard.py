import base64
import io
import re
import textwrap
import unicodedata
from decimal import Decimal, InvalidOperation

import requests
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook

from odoo import _, fields, models
from odoo.exceptions import AccessError, UserError, ValidationError


COMPANY_ID = 10
POS_CONFIG_ID = 16
POS_ROOT_ID = 203
MERCEDES_USER_ID = 13


class JkgMelocotonOpsMixin(models.AbstractModel):
    _name = "jkg.melocoton.ops.mixin"
    _description = "Utilidades protegidas de productos El Melocotón"

    def _check_operator(self):
        if self.env.uid in (1, MERCEDES_USER_ID):
            return
        if self.env.user.has_group("base.group_system"):
            return
        raise AccessError(_("Esta herramienta está reservada a El Melocotón y JKG."))

    @staticmethod
    def _normalize(value):
        value = unicodedata.normalize("NFKD", value or "")
        value = "".join(
            character
            for character in value
            if not unicodedata.combining(character)
        )
        return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()

    @staticmethod
    def _valid_ean13(code):
        if not code:
            return True
        if len(code) != 13 or not code.isdigit():
            return False
        digits = [int(value) for value in code]
        total = sum(digits[0:12:2]) + 3 * sum(digits[1:12:2])
        return (10 - total % 10) % 10 == digits[12]

    def _company(self):
        company = self.env["res.company"].sudo().browse(COMPANY_ID).exists()
        if not company:
            raise UserError(_("No existe la empresa de El Melocotón."))
        return company

    def _product_model(self):
        company = self._company()
        return (
            self.env["product.template"]
            .sudo()
            .with_company(company)
            .with_context(
                active_test=False,
                allowed_company_ids=[COMPANY_ID],
                tracking_disable=True,
                mail_notrack=True,
                jkg_product_import=True,
                jkg_product_guard_bypass=True,
                jkg_allow_product_create=True,
                jkg_melocoton_allow_product_create=True,
                jkg_melocoton_quick_create=True,
            )
        )

    def _find_tax(self, rate):
        company = self._company()
        Tax = self.env["account.tax"].sudo().with_company(company)
        tax = Tax.search(
            [
                ("company_id", "=", COMPANY_ID),
                ("type_tax_use", "=", "sale"),
                ("amount", "=", float(rate)),
                ("active", "=", True),
            ],
            order="id",
            limit=1,
        )
        if not tax:
            raise UserError(_("No se encontró el IVA de venta del %s%%.") % rate)
        return tax

    def _classify(self, name, barcode=None):
        normalized = self._normalize(name)

        rules = [
            (["agua fuerte", "lejia", "limpiador", "detergente"], 21, ["limpieza"]),
            (["energy", "ener drink", "energetica"], 21, ["energeticas", "bebidas"]),
            (["limonada", "refresco", "cola", "gaseosa"], 21, ["bebidas"]),
            (["queso"], 4, ["quesos", "frescos"]),
            (["ruffles", "patatas fritas", "snack", "aperitivo"], 10, ["aperitivos"]),
            (["montadito", "piripi", "pollo en salsa", "preparado infantil"], 10, ["platos preparados"]),
            (["paella", "poton", "congelado"], 10, ["congelados"]),
            (["empanada", "hojaldre", "empanadilla", "pan"], 10, ["panaderia"]),
            (["jamon", "charcuteria", "fiambre"], 10, ["charcuteria", "frescos"]),
            (["tomate frito", "conserva", "salsa"], 10, ["conservas", "despensa"]),
            (["agua"], 10, ["bebidas"]),
            (["platano", "manzana", "pera", "naranja", "mandarina", "limon", "tomate", "lechuga", "cebolla", "patata", "fruta", "verdura"], 4, ["fruteria"]),
        ]

        for keywords, rate, category_aliases in rules:
            if any(keyword in normalized for keyword in keywords):
                sell_by_weight = (
                    any(token in normalized for token in ("kilo", " kg", "al peso"))
                    or (not barcode and (rate == 4 or "fruteria" in category_aliases))
                )
                return rate, category_aliases, sell_by_weight

        sell_by_weight = any(
            token in normalized for token in ("kilo", " kg", "al peso")
        )
        return 10, ["otros"], sell_by_weight

    def _find_or_create_pos_category(self, aliases):
        root = self.env["pos.category"].sudo().browse(POS_ROOT_ID).exists()
        if not root:
            raise UserError(_("No existe la categoría raíz de El Melocotón."))

        categories = self.env["pos.category"].sudo().search(
            [("id", "child_of", root.id)]
        )
        indexed = {}
        for category in categories:
            indexed[self._normalize(category.name)] = category
            indexed[self._normalize(category.display_name)] = category

        for alias in aliases:
            normalized_alias = self._normalize(alias)
            exact = indexed.get(normalized_alias)
            if exact:
                return exact
            for key, category in indexed.items():
                if key.endswith(" " + normalized_alias) or normalized_alias in key.split(" / "):
                    return category

        label = aliases[0].strip().title() if aliases else "Otros"
        return self.env["pos.category"].sudo().create(
            {"name": label, "parent_id": root.id}
        )

    def _find_or_create_internal_category(self, aliases):
        Category = self.env["product.category"].sudo()
        root = Category.search([("name", "=ilike", "El Melocotón")], limit=1)
        if not root:
            root = Category.create({"name": "El Melocotón"})

        children = Category.search([("parent_id", "=", root.id)])
        indexed = {self._normalize(category.name): category for category in children}
        for alias in aliases:
            found = indexed.get(self._normalize(alias))
            if found:
                return found

        label = aliases[0].strip().title() if aliases else "Otros"
        return Category.create({"name": label, "parent_id": root.id})

    def _uom(self, sell_by_weight):
        if sell_by_weight:
            uom = self.env.ref("uom.product_uom_kgm", raise_if_not_found=False)
            if not uom:
                uom = self.env["uom.uom"].sudo().search(
                    [("name", "in", ["kg", "Kilogram", "Kilogramo"])],
                    limit=1,
                )
        else:
            uom = self.env.ref("uom.product_uom_unit", raise_if_not_found=False)
            if not uom:
                uom = self.env["uom.uom"].sudo().search([], limit=1)
        if not uom:
            raise UserError(_("No se encontró la unidad de medida necesaria."))
        return uom

    def _open_food_facts(self, barcode):
        if not barcode:
            return {}
        try:
            response = requests.get(
                "https://world.openfoodfacts.org/api/v2/product/%s.json" % barcode,
                params={
                    "fields": "code,product_name,product_name_es,brands,quantity,image_front_url,image_url"
                },
                headers={
                    "User-Agent": "JKGComputer-OdooProductIntake/1.0 (https://thejoseka.com)"
                },
                timeout=(4, 10),
            )
            if response.status_code != 200:
                return {}
            payload = response.json()
            if payload.get("status") != 1:
                return {}
            return payload.get("product") or {}
        except Exception:
            return {}

    @staticmethod
    def _font(size, bold=False):
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-%s.ttf"
            % ("Bold" if bold else ""),
            "/usr/share/fonts/truetype/liberation2/LiberationSans-%s.ttf"
            % ("Bold" if bold else "Regular"),
        ]
        for path in candidates:
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
        return ImageFont.load_default()

    def _representative_image(self, name, subtitle):
        canvas = Image.new("RGB", (1024, 1024), (248, 248, 244))
        draw = ImageDraw.Draw(canvas)
        draw.rounded_rectangle(
            (64, 64, 960, 960),
            radius=46,
            fill=(255, 255, 255),
            outline=(87, 122, 52),
            width=8,
        )
        draw.ellipse(
            (372, 170, 652, 450),
            fill=(245, 193, 89),
            outline=(87, 122, 52),
            width=7,
        )
        title_font = self._font(50, bold=True)
        subtitle_font = self._font(34)
        note_font = self._font(25, bold=True)
        y = 520
        for line in textwrap.wrap(name, width=24):
            box = draw.textbbox((0, 0), line, font=title_font)
            draw.text(
                ((1024 - (box[2] - box[0])) / 2, y),
                line,
                font=title_font,
                fill=(31, 63, 33),
            )
            y += 66
        box = draw.textbbox((0, 0), subtitle, font=subtitle_font)
        draw.text(
            ((1024 - (box[2] - box[0])) / 2, y + 20),
            subtitle,
            font=subtitle_font,
            fill=(70, 70, 70),
        )
        note = "IMAGEN REPRESENTATIVA"
        box = draw.textbbox((0, 0), note, font=note_font)
        draw.text(
            ((1024 - (box[2] - box[0])) / 2, 865),
            note,
            font=note_font,
            fill=(65, 96, 42),
        )
        output = io.BytesIO()
        canvas.save(output, format="JPEG", quality=88, optimize=True)
        return output.getvalue()

    def _download_image(self, url):
        if not url:
            return None
        try:
            response = requests.get(
                url,
                headers={
                    "User-Agent": "JKGComputer-OdooProductIntake/1.0 (https://thejoseka.com)"
                },
                timeout=(4, 12),
            )
            response.raise_for_status()
            if not response.content or len(response.content) > 6 * 1024 * 1024:
                return None
            with Image.open(io.BytesIO(response.content)) as source:
                source = source.convert("RGB")
                source.thumbnail((920, 920))
                canvas = Image.new("RGB", (1024, 1024), "white")
                canvas.paste(
                    source,
                    ((1024 - source.width) // 2, (1024 - source.height) // 2),
                )
                output = io.BytesIO()
                canvas.save(output, format="JPEG", quality=88, optimize=True)
                return output.getvalue()
        except Exception:
            return None

    def _build_description(self, name, info, sell_by_weight):
        parts = []
        product_name = info.get("product_name_es") or info.get("product_name")
        brand = (info.get("brands") or "").split(",")[0].strip()
        quantity = (info.get("quantity") or "").strip()
        if product_name and self._normalize(product_name) != self._normalize(name):
            parts.append(product_name.strip())
        if brand:
            parts.append("Marca: %s" % brand)
        if quantity:
            parts.append("Formato: %s" % quantity)
        if sell_by_weight:
            parts.append("Venta al peso")
        if not parts:
            parts.append("Producto disponible en El Melocotón")
        return ". ".join(parts) + "."

    def _ensure_no_duplicate(self, name, barcode):
        Product = self._product_model()
        if barcode:
            existing = Product.search([("barcode", "=", barcode)], limit=2)
            if existing:
                raise ValidationError(
                    _("El EAN %s ya pertenece a %s. Use Cambiar precio.")
                    % (barcode, existing[0].display_name)
                )
        existing_name = Product.search(
            [
                ("name", "=ilike", name.strip()),
                ("company_id", "in", [False, COMPANY_ID]),
            ],
            limit=1,
        )
        if existing_name:
            raise ValidationError(
                _("Ya existe un producto llamado %s. Use Cambiar precio.")
                % existing_name.display_name
            )

    def _create_product(self, name, price, barcode=None, operation="create"):
        self._check_operator()
        name = (name or "").strip()
        barcode = (barcode or "").strip() or False
        if not name:
            raise ValidationError(_("El nombre es obligatorio."))
        if not price or float(price) <= 0:
            raise ValidationError(_("El precio debe ser mayor que cero."))
        if barcode and not self._valid_ean13(barcode):
            raise ValidationError(_("El EAN %s no es válido.") % barcode)

        self._ensure_no_duplicate(name, barcode)
        vat_rate, category_aliases, sell_by_weight = self._classify(name, barcode)
        tax = self._find_tax(vat_rate)
        pos_category = self._find_or_create_pos_category(category_aliases)
        internal_category = self._find_or_create_internal_category(category_aliases)
        uom = self._uom(sell_by_weight)
        info = self._open_food_facts(barcode)
        image_url = info.get("image_front_url") or info.get("image_url")
        image_data = self._download_image(image_url)
        image_source = "EAN" if image_data else "REPRESENTATIVA"
        if not image_data:
            image_data = self._representative_image(name, pos_category.display_name)

        values = {
            "name": name,
            "list_price": float(price),
            "company_id": COMPANY_ID,
            "categ_id": internal_category.id,
            "taxes_id": [(6, 0, [tax.id])],
            "active": True,
            "sale_ok": True,
            "purchase_ok": True,
            "uom_id": uom.id,
            "default_code": "MEL-%s" % barcode if barcode else "MEL-SINCOD-%s" % re.sub(r"[^A-Z0-9]+", "-", self._normalize(name).upper()).strip("-")[:28],
            "image_1920": base64.b64encode(image_data),
        }
        Product = self._product_model()
        if "barcode" in Product._fields:
            values["barcode"] = barcode
        if "available_in_pos" in Product._fields:
            values["available_in_pos"] = True
        if "pos_categ_ids" in Product._fields:
            values["pos_categ_ids"] = [(6, 0, [pos_category.id])]
        if "to_weight" in Product._fields:
            values["to_weight"] = sell_by_weight
        if "uom_po_id" in Product._fields:
            values["uom_po_id"] = uom.id
        if "is_storable" in Product._fields:
            values["is_storable"] = True
        if "description_sale" in Product._fields:
            values["description_sale"] = self._build_description(
                name, info, sell_by_weight
            )

        product = Product.create(values)
        self.env["jkg.melocoton.product.operation.log"].sudo().create(
            {
                "operation": operation,
                "product_id": product.id,
                "barcode": barcode,
                "new_price": float(price),
                "status": "ok",
                "detail": _("IVA %s%% · %s · imagen %s")
                % (vat_rate, pos_category.display_name, image_source),
                "company_id": COMPANY_ID,
            }
        )
        return product

    def _notification(self, title, message, notification_type="success"):
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": title,
                "message": message,
                "type": notification_type,
                "sticky": False,
                "next": {"type": "ir.actions.act_window_close"},
            },
        }


class JkgMelocotonNewProductWizard(models.TransientModel):
    _name = "jkg.melocoton.new.product.wizard"
    _inherit = "jkg.melocoton.ops.mixin"
    _description = "Nuevo producto protegido El Melocotón"

    barcode = fields.Char(string="EAN", help="Opcional si el producto no tiene código")
    name = fields.Char(string="Nombre", required=True)
    price = fields.Float(string="Precio de venta", required=True, digits="Product Price")

    def action_create_product(self):
        self.ensure_one()
        product = self._create_product(self.name, self.price, self.barcode)
        return self._notification(
            _("Producto listo"),
            _("%s ya está activo y disponible en el TPV a %.2f €.")
            % (product.display_name, product.list_price),
        )


class JkgMelocotonPriceWizard(models.TransientModel):
    _name = "jkg.melocoton.price.wizard"
    _inherit = "jkg.melocoton.ops.mixin"
    _description = "Cambio protegido de precio El Melocotón"

    product_id = fields.Many2one(
        comodel_name="product.template",
        string="Producto",
        required=True,
        domain="[('company_id', '=', 10), ('active', '=', True), ('sale_ok', '=', True)]",
    )
    current_price = fields.Float(
        string="Precio actual",
        related="product_id.list_price",
        readonly=True,
    )
    new_price = fields.Float(
        string="Nuevo precio",
        required=True,
        digits="Product Price",
    )

    def action_change_price(self):
        self.ensure_one()
        self._check_operator()
        if self.new_price <= 0:
            raise ValidationError(_("El precio debe ser mayor que cero."))
        product = self.product_id.sudo().with_context(
            tracking_disable=True,
            mail_notrack=True,
            jkg_product_guard_bypass=True,
            jkg_allow_product_create=True,
            jkg_melocoton_allow_product_create=True,
        )
        if product.company_id.id != COMPANY_ID:
            raise AccessError(_("Solo se pueden modificar productos de El Melocotón."))
        old_price = product.list_price
        product.write({"list_price": self.new_price})
        self.env["jkg.melocoton.product.operation.log"].sudo().create(
            {
                "operation": "price",
                "product_id": product.id,
                "barcode": product.barcode,
                "old_price": old_price,
                "new_price": self.new_price,
                "status": "ok",
                "detail": _("Cambio protegido de precio"),
                "company_id": COMPANY_ID,
            }
        )
        return self._notification(
            _("Precio actualizado"),
            _("%s: %.2f € → %.2f €")
            % (product.display_name, old_price, self.new_price),
        )


class JkgMelocotonImportWizard(models.TransientModel):
    _name = "jkg.melocoton.import.wizard"
    _inherit = "jkg.melocoton.ops.mixin"
    _description = "Importación protegida Excel El Melocotón"

    file_data = fields.Binary(string="Archivo Excel", required=True, attachment=False)
    file_name = fields.Char(string="Nombre del archivo")
    result_summary = fields.Text(string="Resultado", readonly=True)
    template_file = fields.Binary(string="Plantilla", readonly=True, attachment=False)
    template_filename = fields.Char(readonly=True)

    @staticmethod
    def _header_key(value):
        value = unicodedata.normalize("NFKD", str(value or ""))
        value = "".join(
            character
            for character in value
            if not unicodedata.combining(character)
        )
        return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()

    def action_download_template(self):
        self.ensure_one()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PRODUCTOS"
        sheet.append(["EAN", "Nombre", "Precio"])
        sheet.append(["8436032230276", "Piripi", 1.90])
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 42
        sheet.column_dimensions["C"].width = 15
        output = io.BytesIO()
        workbook.save(output)
        self.write(
            {
                "template_file": base64.b64encode(output.getvalue()),
                "template_filename": "Plantilla_Productos_El_Melocoton.xlsx",
            }
        )
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/?model=%s&id=%s&field=template_file&filename_field=template_filename&download=true"
            % (self._name, self.id),
            "target": "self",
        }

    def action_import(self):
        self.ensure_one()
        self._check_operator()
        if not self.file_data:
            raise ValidationError(_("Seleccione un archivo Excel."))
        try:
            content = base64.b64decode(self.file_data)
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError(_("No se pudo leer el Excel: %s") % exc) from exc

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError(_("El Excel está vacío."))

        headers = {self._header_key(value): index for index, value in enumerate(rows[0])}

        def find_column(options):
            for option in options:
                if option in headers:
                    return headers[option]
            return None

        ean_column = find_column(["ean", "codigo ean", "codigo de barras", "barcode"])
        name_column = find_column(["nombre", "producto", "nombre del articulo", "articulo"])
        price_column = find_column(["precio", "pvp", "precio de venta", "precio venta"])
        if name_column is None or price_column is None:
            raise ValidationError(
                _("El Excel debe contener las columnas Nombre y Precio. EAN es opcional.")
            )

        created = []
        errors = []
        for row_number, row in enumerate(rows[1:], start=2):
            if not row or all(value in (None, "") for value in row):
                continue
            name = row[name_column] if name_column < len(row) else None
            raw_price = row[price_column] if price_column < len(row) else None
            raw_barcode = (
                row[ean_column]
                if ean_column is not None and ean_column < len(row)
                else None
            )
            barcode = ""
            if raw_barcode not in (None, ""):
                if isinstance(raw_barcode, float) and raw_barcode.is_integer():
                    barcode = str(int(raw_barcode))
                else:
                    barcode = str(raw_barcode).strip().replace(".0", "")
            try:
                if isinstance(raw_price, str):
                    raw_price = raw_price.replace("€", "").replace(",", ".").strip()
                price = float(Decimal(str(raw_price)))
            except (InvalidOperation, TypeError, ValueError):
                errors.append(_("Fila %s: precio no válido") % row_number)
                continue
            try:
                with self.env.cr.savepoint():
                    product = self._create_product(
                        str(name or "").strip(),
                        price,
                        barcode,
                        operation="import",
                    )
                    created.append(product.display_name)
            except Exception as exc:
                errors.append(_("Fila %s: %s") % (row_number, str(exc)))

        if not created and errors:
            raise ValidationError("\n".join(errors[:15]))

        summary = _("Creados: %s") % len(created)
        if errors:
            summary += _(" · Revisión: %s") % len(errors)
            summary += "\n" + "\n".join(errors[:20])
        self.result_summary = summary
        return self._notification(
            _("Importación terminada"),
            summary,
            "warning" if errors else "success",
        )
